
import pandas as pd
from openpyxl import *
import xlrd
from bs4 import BeautifulSoup as soup
import requests
import time


# Get text of html
def getText(url, user_agent):
    headers = {"User-Agent": user_agent}
    page = requests.get(url,headers=headers)
    content = soup(page.content, "html.parser")
    for script in content(["script", "style"]):
        script.decompose()
    strip = list(content.stripped_strings)
    for x in range(len(strip)):
        strip[x]=strip[x].lower()
    names=[]
    for link in content.findAll('a'):
        names.append(link.get('href'))
    return(strip + names)

def getText2(url):
    page = requests.get(url)
    content = soup(page.content, "html.parser")
    for script in content(["script", "style"]):
        script.decompose()
    strip = list(content.stripped_strings)
    for x in range(len(strip)):
        strip[x]=strip[x].lower()
    names=[]
    for link in content.findAll('a'):
        names.append(link.get('href'))
    return(strip + names)


# Getting URL for "contact" website
def getURL(url, user_agent):
    headers = {"User-Agent": user_agent}
    page = requests.get(url,headers=headers)
    content = soup(page.content, "html.parser")
    i = 0
    short_url=""
    num=0
    while(i<len(url) and num<3):
        short_url+=url[i]
        if(url[i] == "/"):
            num+=1
        i+=1
    for link in (content.findAll('a',  href = True, text= ('Contact Us' or 'Patient Portal' or 'Patient Portal'))):
        string = link.get('href')
        if(string[:4] == "http"):
            return string
        elif(string[0] == "/" and short_url[len(short_url)-1] == '/'):
            new_string = string.replace("/", "", 1)
            return(short_url+(new_string))
        elif(string[0] != "/" and short_url[len(short_url)-1] != '/'):
            return(short_url + "/" + string)
    return(url)


def getURL2(url, user_agent):
    headers = {"User-Agent": user_agent}
    page = requests.get(url, headers=headers)
    content = soup(page.content, "html.parser")
    names = []
    for link in content.findAll('a'):
        names.append(link.get('href'))
    i = 0
    while(i<len(names)):
        if(str(names[i]).find("contact")>0):
            if((str(names[i]))[0:4]=="http"):
                return(str(names[i]))
            else:
                return(url+names[i])
        i+=1
    return(url)

def getURL3(url):
    page = requests.get(url)
    content = soup(page.content, "html.parser")
    i = 0
    short_url=""
    num=0
    while(i<len(url) and num<3):
        short_url+=url[i]
        if(url[i] == "/"):
            num+=1
        i+=1
    for link in (content.findAll('a',  href = True, text= 'Contact Us')):
        string = link.get('href')
        if(string[:4] == "http"):
            return string
        return(short_url+(link.get('href')))
    return(url)

# Check email
def checkMail(strip):
    i = 0
    while(i<len(strip)):
        if(strip[i] == None):
            i+=1
        elif("@" in strip[i] and ".com" in strip[i]):
            return("1")
        else:
            i+=1
    return("0")

# Check Phone Number
def checkPhone(strip): 
    i = 0
    while(i<len(strip)):
        new_string = strip[i].replace(" ","")
        if(strip[i] == None):
            i+=1
        elif("tel" in strip[i]):
            return("1")
        elif(len(new_string)>10):
            num = len(new_string)-5
            if(new_string[0]=="(" and new_string[4]==")" and new_string[num]=="-"):
                return("1")
            elif(new_string[3]=="-"and new_string[num]=="-"):
                return("1")
            elif(new_string[num]=="-"):
                return("1")
            else:
                i+=1
        else:
            i+=1
    for x in strip:
        if(x.find("phone")>-1):
            return("1")
    return("0")
    
# Check Email Form
def checkForm(strip):
    i = 0
    while(i<len(strip)):
        if(strip[i] == None):
            i+=1
        elif((strip[i].find(" form")>-1) or (strip[i].find(" submit")>-1)):
            return("1")
        else:
            i+=1
    return("0")

# Check TDD TYY 
def checkT(strip):
    i = 0
    while(i<len(strip)):
        if(strip[i] == None):
            i+=1
        elif((strip[i].find("tdd")>-1) or (strip[i].find("tty")>-1)):
            return("1")
        else:
            i+=1
    return("0")

# Check Text Message
def checkText(strip):
    i = 0
    while(i<len(strip)):
        if(strip[i] == None):
            i+=1
        elif((strip[i].find("text message")>-1) or (strip[i].find("text messaginng")>-1) or (strip[i].find("chat")>-1)):
            return("1")
        else:
            i+=1
    return("0")

# Check Text Relay
def checkRelay(strip):
    i = 0
    while(i<len(strip)):
        if(strip[i] == None):
            i+=1
        elif((strip[i].find("zoom")>-1) or (strip[i].find("skype")>-1) or
         (strip[i].find("video conferencing")>-1) or (strip[i].find("video chat")>-1)):
            return("1")
        else:
            i+=1
    return("0")

print("Start")


start = time.time()
# Basic information of the excel worksheet
excel1 = "hospital.xlsx"
d1 = pd.read_excel(excel1)
workbook = xlrd.open_workbook("hospital.xlsx")
sheet = workbook.sheet_by_index(0)
book = load_workbook("hospital.xlsx") # For editing excel
page = book["Sheet1"]

# Creating list of websites
names=[]
website_list=[]
website_to_list = {}
i = 3 
for i in range(sheet.nrows):
    website_list.append(sheet.cell_value(i,15))
    names.append(sheet.cell_value(i,2))

    # Creating a dictionary to get the name of each websites
    website_to_list[sheet.cell_value(i,15)] = sheet.cell_value(i,2)
    i+=1
# print(website_to_list)
print(website_list)






# Getting URL and information of a website   
# Checking each websites
google = "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)"
personal = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.67 Safari/537.36 Edg/87.0.664.47"
num_badWebsites, total_phone, total_email, total_form, total_relay, total_t, total_text = 0,0,0,0,0,0,0
start1=time.time()

list_phone = []
list_mail = []
list_form = []
list_relay = []
list_tty = []
list_text = []
bad_websites = []

list_phone_name = []
list_mail_name = []
list_form_name = []
list_relay_name = []
list_tty_name = []
list_text_name = []
bad_websites_name = []


good_list_phone = []
good_list_mail = []
good_list_form = []
good_list_relay = []
good_list_tty = []
good_list_text = []

good_list_phone_name = []
good_list_mail_name = []
good_list_form_name = []
good_list_relay_name = []
good_list_tty_name = []
good_list_text_name = []


blocked = []
forbidden = []
default = []
not_found = []
cannot_decode = []
javascript = []

blocked_name = []
forbidden_name = []
default_name = []
not_found_name = []
cannot_decode_name = []
javascript_name = []

w = 3
while(w<len(website_list)):
    list1=[]
    start= time.time()
    try:        
        print(website_list[w])
        if(getURL(website_list[w],personal) != None):
            url = getURL(website_list[w],personal)  

        elif(getURL(website_list[w],google) != None):
            url = getURL(website_list[w],google)

        elif(getURL2(website_list[w],google) != None):
            url = getURL2(website_list[w],google)

        elif(getText2(website_list[w]).count('403 forbidden') == 0):
            url = getURL3(website_list[w])
            
        else:
            url = website_list[w]
    except:
        pass
    end= time.time()
    print(f"getURL uses {end-start} seconds")
    print()
    try:
        info = getText(url, personal)
        if(info.count("403 forbidden") > 0):
            info = getText(url, google)
            if(info[0]==("attention required! | cloudflare")):
                info = getText2(url)
        error=0
        symbol=0
        java=0
        for x in info:
            try:
                if(x.count("404" or "not found")>0):
                    error+=1
                if(x.count("�")>0):
                    symbol+=1
                if(x.count("javascript is disabled")>0):
                    java+=1
            except:
                pass
        if(error>0):
            not_found.append(website_list[w])
            # not_found_name.append(names[w])
        elif(symbol>0):
            cannot_decode.append(website_list[w])
            # cannot_decode_name.append(names[w])
        elif(java>0):
            javascript.append(website_list[w])
            # javascript_name.append(names[w])
        elif(info[0]=="Web Page Blocked"):
            blocked.append(website_list[w])
            # blocked_name.append(names[w])
        elif(info.count("403 forbidden")>0):
            forbidden.append(website_list[w])
            # forbidden_name.append(names[w])
        elif(info[0]== "default web site page"):
            default.append(website_list[w])
            # default_name.append(names[w])
        else:
            # if(info[0]!="Web Page Blocked" and info.count("403 forbidden")==0 and info[0]!= "default web site page"):
            page.cell(w+1,8).value=checkPhone(info)
            if(checkPhone(info) == "1"):
                total_phone+=1
                good_list_phone.append(website_list[w])
                # good_list_phone_name.append(names[w])
            else:
                list_phone.append(website_list[w])
                # list_phone_name.append(names[w])

            page.cell(w+1,9).value=checkMail(info)
            if(checkMail(info) == "1"):
                total_email+=1
                good_list_mail.append(website_list[w])
                # good_list_mail_name.append(names[w])
            else:
                list_mail.append(website_list[w])
                # list_mail_name.append(names[w])

            page.cell(w+1,10).value=checkForm(info)
            if(checkForm(info) == "1"):
                total_form+=1
                good_list_form.append(website_list[w])
                # good_list_form_name.append(names[w])
            else:
                list_form.append(website_list[w])
                # list_form_name.append(names[w])

            page.cell(w+1,11).value=checkT(info)
            if(checkT(info) == "1"):
                total_t+=1
                good_list_tty.append(website_list[w])
                # good_list_tty_name.append(names[w])
            else:
                list_tty.append(website_list[w])
                # list_tty_name.append(names[w])

            page.cell(w+1,12).value=checkText(info)
            if(checkText(info) == "1"):
                total_text+=1
                good_list_text.append(website_list[w])
                # good_list_text_name.append(names[w])
            else:
                list_text.append(website_list[w])
                # list_text_name.append(names[w])

            page.cell(w+1,13).value=checkRelay(info)
            if(checkRelay(info) == "1"):
                total_relay+=1
                good_list_relay.append(website_list[w])
                # good_list_relay_name.append(names[w])
            else:
                list_relay.append(website_list[w])
                # list_relay_name.append(names[w])
            page.cell(w+1,14).value=checkT(info)

            
    except:
        page.cell(w+1,8).value=(str(website_list[w]) + "Not Good")
        print("Line "+ str(w) +" is not good") 
        num_badWebsites += 1
        bad_websites.append(website_list[w])
    w+=1

book.save("hospital.xlsx")

# printing out console report
print("There is " +str(total_phone) + " Phone")
print("There is " +str(total_email) + " email")
print("There is " +str(total_form) + " form")
print("There is " +str(total_text) + " text")
print("There is " +str(total_t) + " TTY")
print("There is " +str(total_relay) + " relay")
print("There is " +str(num_badWebsites) + " unusable websites")



# Print all the websites that have issue
print(f"Name Blocked {blocked_name}")
print("Break")
print(f"Name Forbidden {forbidden_name}")
print("Break")
print(f"Name default {default_name}")
print("Break")
print(f"Name not_found {not_found_name}")
print("Break")
print(f"Name Cant Decode {cannot_decode_name}")
print("Break")
print(f"Name Java Script {javascript_name}")
print("Break")
input()



